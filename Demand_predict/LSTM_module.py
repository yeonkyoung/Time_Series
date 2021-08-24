#-*-coding:utf-8-*-
import pandas as pd
import numpy as np
from pandas import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import time
import tensorflow as tf
from datetime import timedelta
import autokeras as ak
from kerastuner import HyperModel
from sklearn.preprocessing import StandardScaler,MinMaxScaler,RobustScaler
from keras.models import Sequential
from keras.layers import Dense, LSTM, Dropout
from keras.models import load_model
from pmdarima import model_selection
from openpyxl import Workbook, drawing

def result_input(wb,model_root,test_h,col,com_col,output,pre_graph,loss_graph):
    sheet1 = wb.active
    sheet1.cell(row=1, column=2).value = col
    sheet1.cell(row=2, column=2).value = model_root
    sheet1.cell(row=4, column=1).value = output.index[0]
    sheet1.cell(row=3, column=2).value = col
    sheet1.cell(row=4, column=2).value = output[col][0]
    sheet1.cell(row=3, column=3).value = 'prediction'
    sheet1.cell(row=4, column=3).value = output['prediction'][0]
    sheet1.cell(row=3, column=4).value = com_col
    sheet1.cell(row=4, column=4).value = output[com_col][0]
    #graph import
    loss=drawing.image.Image(loss_graph)
    sheet1.add_image(loss,'A7')
    pre=drawing.image.Image(pre_graph)
    sheet1.add_image(pre,'A20')
    return sheet1


#data to timeseries data 
def data2tsdata(raw_data,col,use_date_max):  #column은 '' 사용하여 입력
    
#     data['MM']=data['MONTH'].str.replace('월','')
    raw_data['MONTH']=raw_data['MONTH'].astype(str)
    raw_data['YEAR']=raw_data['YEAR'].astype(str)
    raw_data['Time']=raw_data['YEAR']+raw_data['MONTH'] #data['MM']
    raw_data['Time']=raw_data['Time'].apply(lambda x: datetime.strptime(x, '%Y%m'))
    ts=raw_data[['Time',col]]
    ts_data=ts.set_index('Time')
    ts_datas=ts_data[(ts_data.index<=use_date_max)]

    return ts_datas.dropna()


# scalers
def standardscale(ts_data):
    scal = StandardScaler()
    dd_t = scal.fit_transform(ts_data)
    data_t = pd.DataFrame(dd_t,index=ts_data.index)
    return (data_t,scal)

def minmaxscale(ts_data):
    scal = MinMaxScaler()
    dd_t = scal.fit_transform(ts_data)
    data_t = pd.DataFrame(dd_t,index=ts_data.index)
    return (data_t,scal)
    
def robustscale(ts_data):
    scal = RobustScaler()
    dd_t = scal.fit_transform(ts_data)
    data_t = pd.DataFrame(dd_t,index=ts_data.index)
    return (data_t,scal)


#train_test split
def train_test_split(data_t,test_h,n_steps): #,val_h 
#     data=data_t[(data_t.index<=use_date_max)]
    test_cutoff_date= data_t.index.max() - timedelta(days=test_h*n_steps*28) #h = 1
#     val_cutoff_date=data_t.index.max() - timedelta(days=2*val_h*30) 
    df_test = data_t[data_t.index >=  test_cutoff_date ]
#     df_val = data_t[(data_t.index > val_cutoff_date) & (data_t.index <= test_cutoff_date)]
    df_train = data_t[(data_t.index < test_cutoff_date)]  #val_cutoff_date
    return (df_train,df_test) #,df_val


#preprocessing of LSTM
def split_sequence(sequence, n_steps:int):  #:ndarray   ,   ->(ndarray, ndarray)
    x = []
    y = [] 
    for i in range(len(sequence)):
        if(i+n_steps>=len(sequence)):
            break 
        x.append(sequence[i:i+n_steps]) 
        y.append(sequence[i+n_steps]) 
    return (np.asarray(x), np.asarray(y)) 



# # build model
# # class LSTM_model() :
# n_steps=5
# simple_lstm_model = tf.keras.models.Sequential([
# tf.keras.layers.LSTM(10, activation = 'relu', input_shape=(n_steps,1),return_sequences=True),
# tf.keras.layers.Dropout(0.2),
# tf.keras.layers.LSTM(20, activation = 'relu', input_shape=(n_steps,1),return_sequences=True),
# tf.keras.layers.Dropout(0.2),
# tf.keras.layers.LSTM(10, activation = 'relu', input_shape=(n_steps,1)), 
# tf.keras.layers.Dropout(0.2),
# tf.keras.layers.Dense(5),   
# tf.keras.layers.Dense(1)])


def visualize_loss(history, title,saveroot):
    loss = history.history["loss"]
    val_loss = history.history["val_loss"]
    epochs = range(len(loss))
    plt.plot(epochs, loss, "b", label="Training loss")
    plt.plot(epochs, val_loss, "r", label="Validation loss")
    plt.title(title)
    plt.xlabel("Epochs")
    plt.ylabel("Loss")
    plt.legend()
    plt.savefig(saveroot+'lossgraph.png')
    return plt.show()



def predict_graph(train,test,prediction,plan,saveroot) :
    plt.figure(figsize=(8,5))
    kws = dict(marker='o')
    plt.plot(train, label='Train', **kws)   
    
    plt.plot(test, label='Test', **kws)   
    plt.plot(prediction,label='predicted', **kws)
    plt.plot(plan,label='plan', **kws)
    plt.legend()
    plt.savefig(saveroot+'_timeseries.png')
    return plt.show()



##################MAIN############################

#모델 학습(LSTM) & 평가
def LSTM_uni_train(raw_data,use_date_max,col,com_col,scale,n_steps,n_features,test_h,model,BATCH_SIZE=1,BUFFER_SIZE=100,EVALUATION_INTERVAL = 100,EPOCHS = 1000,optimizer='adam',loss='mse',metrics=['mse'],saveroot='C:/Users/KIMYEONKYOUNG/Desktop/2021 AI 빅데이터팀/메탈 수요예측/code/model_회사별/'):  #scaleoption : None,standard,minmax,robust # n_features = 1 (univariate) 
    
    #raw data 2 time series data
    ts_data=data2tsdata(raw_data,col,use_date_max)
    
    #scaling
    if scale == None :
        data_t = ts_data
    if scale == 'standard':
        (data_t,scal) = standardscale(ts_data)
    if scale == 'minmax':
        (data_t,scal) = minmaxscale(ts_data)
    if scale == 'robust':
        (data_t,scal) = robustscale(ts_data)
    
    #get train data
    (df_train,df_test) = train_test_split(data_t,test_h,n_steps)
    # [train] dateaframe to tensor 
    nd_train=np.asarray(df_train)
    nd_train=nd_train.reshape(len(df_train),)
    nd_test=np.asarray(df_test)
    nd_test=nd_test.reshape(len(df_test),)
    
    (train_x, train_y) = split_sequence(nd_train, n_steps) 
    train_x = train_x.reshape(train_x.shape[0], n_steps, n_features)
    
    #train data 2 train & val
    train_univariate = tf.data.Dataset.from_tensor_slices((train_x, train_y))
    train_univariate = train_univariate.cache().shuffle(BUFFER_SIZE).batch(BATCH_SIZE).repeat()
    val_univariate = tf.data.Dataset.from_tensor_slices((train_x, train_y))
    val_univariate = val_univariate.batch(BATCH_SIZE).repeat()
    
    # Build EarlyStopping
    path_checkpoint = "lstm_model_checkpoint_try.h5"
    es_callback = tf.keras.callbacks.EarlyStopping(monitor="loss", min_delta=0, patience=100, mode='auto')# mode=auto loss면 최저값100번정도 반복되면 정지, acc면 최고값이 100번정도 반복되면 정지
    modelckpt_callback = tf.keras.callbacks.ModelCheckpoint(monitor="loss",filepath=path_checkpoint,verbose=1,save_weights_only=True,save_best_only=True,)
    
    model.compile(optimizer, loss , metrics)
    
    #train model
    history=model.fit(train_univariate,epochs=EPOCHS,validation_data=val_univariate,steps_per_epoch=EVALUATION_INTERVAL,validation_steps=1,verbose=1,callbacks=[es_callback, modelckpt_callback])
    
    #save model
    file_root=col.replace('_실적',"")
    model.save(saveroot+'/'+str(file_root)+'/'+str(use_date_max)+'_'+'lstm_model_checkpoint_'+str(scale)+'_'+str(col)+'.h5')
    
    #graph
    visualize_loss(history, "Training & vaildation Loss",saveroot=saveroot)

    #########################################################################################
    #predict
    
    if test_h == 1 :
        xtt = nd_test.reshape(1, n_steps, n_features)
        #predict
        yhat = model.predict(xtt)
        prediction=pd.DataFrame(yhat)
        prediction.columns=['yhat']
        prediction.index=df_test[4:].index   
    
    if test_h != 1 :
        (xt, yt) = split_sequence(nd_test, n_steps)
        xtt = xt.reshape(xt.shape[0], n_steps, n_features)
        #predict
        yhat = model.predict(xtt)
        prediction=pd.DataFrame(yhat)
        prediction.columns=['yhat']
        prediction.index=df_test[n_step:].index   
    
    #inverse_scale
    prediction['prediction']=scal.inverse_transform(prediction)
    
    train_g, test_g = model_selection.train_test_split(ts_data, train_size=len(ts_data)-test_h)
    #outputdataframe(이동,실적,pred)
    outpu=pd.merge(test_g,prediction['prediction'],left_index=True, right_index=True)
    com_ts_data=data2tsdata(raw_data,com_col,use_date_max)
    output=pd.merge(outpu,com_ts_data,left_index=True, right_index=True)

    #timeseries graph(train,test,predict)    
    predict_graph(train_g,test_g,prediction['prediction'],output[com_col],saveroot=saveroot)

    #excel에 output저장(model_root, test_h, test_train loss, timeseries graph, output df)
    wb = Workbook()
    result_df=result_input(wb,model_root=saveroot+'lstm_model_checkpoint_'+str(scale)+'_'+str(col)+'.h5',test_h=test_h,col=col,com_col=com_col,output=output,pre_graph=saveroot+'_timeseries.png',loss_graph=saveroot+'lossgraph.png')
    wb.save(saveroot+'/'+str(file_root)+'/'+use_date_max+'.xlsx')

    return result_df
